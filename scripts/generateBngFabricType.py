"""
Generates src/data/bngFabricType.json

Parses "BRAS DATA_1.xlsx" (repo root) into a flat list of per-BNG-node
Fabric/Non-Fabric classification records, used as the default snapshot for
the "By Type" tab in BngScreen.tsx (before any upload via /api/bng/upload-fabric).

Primary source: the "Input sheet " tab (HostName, Fabric\\Non Fabric, Circle,
Site, BNG Type, Type, Current Status, Services, ...).

Gap-fill: the "MSAN-BNG" tab's BNG / BNG type columns supply a Fabric/Non-Fabric
tag (verified 100% consistent with Input sheet everywhere the two overlap) for
any node name not already present in Input sheet.

Shutdown status: the "MSC DC status" tab names BNG sites slated for shutdown
(column B = site name, column C = non-empty when that site is "to be shut").
Real site rows have an empty column A; a few stray "new DC under deployment"
reference rows have column A populated and are skipped. Every record's `site`
(Input sheet's Site column / MSAN-BNG's BNG site column) is checked against
this set to set `shuttingDown`.

Run manually: python3 scripts/generateBngFabricType.py
"""

import json
import re
import openpyxl

XLSX_PATH = "BRAS DATA_1.xlsx"
OUT_PATH = "src/data/bngFabricType.json"


def norm_services(raw):
    if raw is None:
        return None
    s = re.sub(r"\s+", " ", str(raw).strip())
    return s or None


def load_shut_sites(wb):
    ws_dc = wb["MSC DC status"]
    shut_sites = set()
    for row in ws_dc.iter_rows(min_row=2, values_only=True):
        site_owner, site, shut_flag = row[0], row[1], row[2]
        if site_owner is not None or not site:
            continue  # stray "new DC under deployment" reference row, not a site row
        if shut_flag:
            shut_sites.add(str(site).strip())
    return shut_sites


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    shut_sites = load_shut_sites(wb)
    print(f"Sites flagged to be shut: {sorted(shut_sites)}")

    input_sheet_key = next(name for name in wb.sheetnames if name.strip() == "Input sheet")
    ws_input = wb[input_sheet_key]
    records = {}
    for row in ws_input.iter_rows(min_row=2, values_only=True):
        host = row[0]
        if not host or str(host).strip() in ("#N/A", "N/A", "#REF!"):
            continue
        node = str(host).strip()
        fabric_raw = str(row[1]).strip() if row[1] is not None else None
        fabric_type = fabric_raw if fabric_raw in ("Fabric", "Non Fabric") else "Unknown"
        site = str(row[3]).strip() if row[3] is not None else None
        records[node] = {
            "node": node,
            "fabricType": fabric_type,
            "bngType": (str(row[4]).strip() or None) if row[4] is not None else None,
            "status": (str(row[6]).strip() or None) if row[6] is not None else None,
            "services": norm_services(row[7]),
            "site": site,
            "shuttingDown": site in shut_sites if site else False,
        }

    ws_msan = wb["MSAN-BNG"]
    for row in ws_msan.iter_rows(min_row=2, values_only=True):
        bng = row[5]
        site = row[7]
        btype = row[8]
        if not bng or str(bng).strip() in ("#N/A", "N/A", "#REF!"):
            continue
        node = str(bng).strip()
        if node in records:
            continue
        fabric_type = btype if btype in ("Fabric", "Non Fabric") else "Unknown"
        site_trimmed = str(site).strip() if site is not None else None
        records[node] = {
            "node": node,
            "fabricType": fabric_type,
            "bngType": None,
            "status": None,
            "services": None,
            "site": site_trimmed,
            "shuttingDown": site_trimmed in shut_sites if site_trimmed else False,
        }

    out = sorted(records.values(), key=lambda r: r["node"])
    with open(OUT_PATH, "w") as f:
        json.dump(out, f, indent=2)

    print(f"Wrote {len(out)} records to {OUT_PATH}")
    fabric = sum(1 for r in out if r["fabricType"] == "Fabric")
    non_fabric = sum(1 for r in out if r["fabricType"] == "Non Fabric")
    unknown = sum(1 for r in out if r["fabricType"] == "Unknown")
    shutting = sum(1 for r in out if r["shuttingDown"])
    print(f"Fabric: {fabric}, Non Fabric: {non_fabric}, Unknown: {unknown}")
    print(f"Shutting down: {shutting} node(s)")


if __name__ == "__main__":
    main()
